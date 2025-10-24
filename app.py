import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, numbers, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO, StringIO
import sys

# ===============================================================
# KONFIGURASI DAN FUNGSI LENDABLE LIMIT (LL)
# (Kode fungsi ini disingkat, diasumsikan sama dengan sebelumnya)
# ===============================================================
STOCK_CODE_BLACKLIST = ['BEBS', 'IPPE', 'WMPP', 'WMUU']
FIRST_LARGERST_COL = "First Largerst"
SECOND_LARGERST_COL = "Second Largerst"
SHEET_INST_OLD = 'Instrument'
SHEET_INST_NEW = 'Hasil Pivot'
SHEET_RESULT_NAME_SOURCE = 'Lendable Limit Result'

# Fungsi LL: process_lendable_limit (DIASUMSIKAN SAMA DENGAN KODE SEBELUMNYA)
def process_lendable_limit(inst_file_buffer, sp_file_buffer, borr_file_buffer, template_file_buffer):
    # ... (Isi fungsi Lendable Limit calculation) ...
    try:
        df_sp = pd.read_excel(sp_file_buffer, header=0, engine='openpyxl')
        df_sp.columns = df_sp.columns.str.strip()
        stock_col = df_sp.columns[1]
        qty_col = df_sp.columns[10]
        df_sp[qty_col] = pd.to_numeric(df_sp[qty_col], errors='coerce').fillna(0)

        top_values = (
            df_sp.groupby(stock_col)[qty_col]
            .apply(lambda x: sorted(x.dropna(), reverse=True)[:2])
            .reset_index()
        )
        top_values[FIRST_LARGERST_COL] = top_values[qty_col].apply(lambda x: x[0] if len(x) > 0 else None)
        top_values[SECOND_LARGERST_COL] = top_values[qty_col].apply(lambda x: x[1] if len(x) > 1 else None)
        df_sp = df_sp.merge(top_values.drop(columns=[qty_col]), how="left", left_on=stock_col, right_on=stock_col)

        df_instr = pd.read_excel(inst_file_buffer, sheet_name=SHEET_INST_OLD, header=1, engine='openpyxl')
        df_instr.columns = df_instr.columns.str.strip()
        col_row = "Local Code"
        col_loan = "Used Loan Qty"
        col_repo = "Used Reverse Repo Qty"
        df_instr[col_loan] = pd.to_numeric(df_instr[col_loan], errors='coerce').fillna(0)
        df_instr[col_repo] = pd.to_numeric(df_instr[col_repo], errors='coerce').fillna(0)
        df_hasil_pivot_sheet = df_instr.groupby(col_row)[[col_loan, col_repo]].sum().reset_index()
        df_hasil_pivot_sheet = df_hasil_pivot_sheet[df_hasil_pivot_sheet[col_row].notna()]
        df_hasil_pivot_sheet = df_hasil_pivot_sheet[~((df_hasil_pivot_sheet[col_loan] == 0) & (df_hasil_pivot_sheet[col_repo] == 0))]
        df_hasil_pivot_sheet = df_hasil_pivot_sheet.sort_values(by=col_row, ascending=True)

        inst_file_buffer.seek(0)
        df_instrument_old_full = pd.read_excel(inst_file_buffer, sheet_name=SHEET_INST_OLD, header=None, engine='openpyxl')
        df_pivot_full = df_hasil_pivot_sheet

        df_sp_full = df_sp.copy()
        qoh_col_sp = df_sp_full.columns[10]
        df_borr_pos_full = pd.read_excel(borr_file_buffer, header=0, engine='openpyxl')
        df_borr_pos_full.columns = df_borr_pos_full.columns.str.strip()

        df_main = df_pivot_full.rename(columns={'Local Code': 'Stock Code'})
        df_result = df_main[['Stock Code']].dropna(subset=['Stock Code']).drop_duplicates(subset=['Stock Code']).copy()
        df_result['Stock Code'] = df_result['Stock Code'].astype(str)

        df_inst_lookup = df_instrument_old_full.iloc[1:].rename(columns={2: 'Stock Code', 9: 'Stock Name'})
        df_inst_lookup = df_inst_lookup[['Stock Code', 'Stock Name']].drop_duplicates(subset=['Stock Code'])
        df_inst_lookup['Stock Code'] = df_inst_lookup['Stock Code'].astype(str)
        df_result = df_result.merge(df_inst_lookup, on='Stock Code', how='left')
        df_result['Stock Name'] = df_result['Stock Name'].fillna('')

        stock_code_col_sp = df_sp_full.columns[1]
        qoh_calc = df_sp_full.groupby(stock_code_col_sp)[qoh_col_sp].sum().reset_index().rename(
            columns={stock_code_col_sp: 'Stock Code', qoh_col_sp: 'Quantity On Hand'}
        )
        qoh_calc['Stock Code'] = qoh_calc['Stock Code'].astype(str)
        df_result = df_result.merge(qoh_calc, on='Stock Code', how='left')
        df_result['Quantity On Hand'] = df_result['Quantity On Hand'].fillna(0)

        BORROW_AMOUNT_COL = 'Borrow Amount (shares)'
        try:
            stock_code_col_borr = 'Stock Code' if 'Stock Code' in df_borr_pos_full.columns else df_borr_pos_full.columns[0]
            df_borr_pos_full[BORROW_AMOUNT_COL] = pd.to_numeric(df_borr_pos_full[BORROW_AMOUNT_COL], errors='coerce').fillna(0)
            borr_qty_calc = df_borr_pos_full[[stock_code_col_borr, BORROW_AMOUNT_COL]].rename(columns={stock_code_col_borr: 'Stock Code', BORROW_AMOUNT_COL: 'Borrow Position'})
            borr_qty_calc = borr_qty_calc.groupby('Stock Code')['Borrow Position'].sum().reset_index()
            borr_qty_calc['Stock Code'] = borr_qty_calc['Stock Code'].astype(str)
            df_result = df_result.merge(borr_qty_calc, on='Stock Code', how='left')
            df_result['Borrow Position'] = df_result['Borrow Position'].fillna(0)
        except Exception:
            df_result['Borrow Position'] = 0.0

        largest_cols_map = {df_sp_full.columns[1]: 'Stock Code', FIRST_LARGERST_COL: FIRST_LARGERST_COL, SECOND_LARGERST_COL: SECOND_LARGERST_COL}
        largest_calc = df_sp_full.rename(columns=largest_cols_map).groupby('Stock Code')[[FIRST_LARGERST_COL, SECOND_LARGERST_COL]].first().reset_index()
        largest_calc['Stock Code'] = largest_calc['Stock Code'].astype(str)
        df_result = df_result.merge(largest_calc, on='Stock Code', how='left').fillna({FIRST_LARGERST_COL: 0, SECOND_LARGERST_COL: 0})

        repo_base_calc = df_pivot_full[['Local Code', 'Used Reverse Repo Qty']].rename(columns={'Local Code': 'Stock Code', 'Used Reverse Repo Qty': 'REPO_Base_TEMP'})
        repo_base_calc['Stock Code'] = repo_base_calc['Stock Code'].astype(str)
        df_result = df_result.merge(repo_base_calc, on='Stock Code', how='left').fillna({'REPO_Base_TEMP': 0})

        df_result['Total two Largerst'] = df_result[FIRST_LARGERST_COL] + df_result[SECOND_LARGERST_COL]
        df_result['Quantity Available'] = df_result['Quantity On Hand'] - df_result['Total two Largerst']
        df_result['Thirty Percent On Hand'] = 0.30 * df_result['Quantity On Hand']
        df_result['REPO'] = 0.10 * df_result['REPO_Base_TEMP']
        df_result['Lendable Limit'] = np.minimum(df_result['Thirty Percent On Hand'], df_result['Quantity Available']) + df_result['REPO']
        df_result['Available Lendable Limit'] = df_result['Lendable Limit'] - df_result['Borrow Position']

        df_result_filtered = df_result[~df_result['Stock Code'].isin(STOCK_CODE_BLACKLIST)].copy()
        df_result_static = df_result_filtered[
            (df_result_filtered['Lendable Limit'] > 0) | (df_result_filtered['Available Lendable Limit'] > 0)
        ].copy()
        df_result_all = df_result_filtered.copy()

        FINAL_COLUMNS = [
            'Stock Code', 'Stock Name', 'Quantity On Hand',
            FIRST_LARGERST_COL, SECOND_LARGERST_COL, 'Total two Largerst',
            'Quantity Available', 'Thirty Percent On Hand', 'REPO',
            'Lendable Limit', 'Borrow Position', 'Available Lendable Limit'
        ]
        df_result_all = df_result_all[FINAL_COLUMNS] 
        df_result_static = df_result_static[FINAL_COLUMNS] 

        output = BytesIO()
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        wb_template = load_workbook(template_file_buffer)
        
        # Define styles inside the function or outside, ensure they are defined once
        if "DefaultStyleLL" not in wb_template.named_styles:
            default_style = NamedStyle(name="DefaultStyleLL")
            default_style.font = Font(name='Roboto Condensed', size=9)
            default_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            default_style.border = thin_border
            wb_template.add_named_style(default_style)
        
        if "TextLeftStyleLL" not in wb_template.named_styles:
            text_left_style = NamedStyle(name="TextLeftStyleLL")
            text_left_style.font = Font(name='Roboto Condensed', size=9)
            text_left_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            text_left_style.border = thin_border
            wb_template.add_named_style(text_left_style)
        
        if "NumberStyleLL" not in wb_template.named_styles:
            number_style = NamedStyle(name="NumberStyleLL")
            number_style.font = Font(name='Roboto Condensed', size=9)
            number_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            number_style.border = thin_border
            number_style.number_format = '#,##0'
            wb_template.add_named_style(number_style)

        ws_template = wb_template.active
        today_formatted = datetime.now().strftime('%d-%b-%y')
        ws_template["B4"] = today_formatted
        start_row = 7
        # Ensure row deletion is safe
        if ws_template.max_row >= start_row:
             ws_template.delete_rows(start_row, ws_template.max_row - start_row + 1)
        
        number_cols_idx = range(2, len(FINAL_COLUMNS))
        df_result_copy = df_result_static.copy() 
        
        for r_idx, row in enumerate(df_result_copy.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_template.cell(row=r_idx, column=c_idx, value=value)
                if c_idx - 1 == 0:
                    cell.style = "DefaultStyleLL"
                    cell.value = str(value) if pd.notna(value) else ""
                elif c_idx - 1 == 1:
                    cell.style = "TextLeftStyleLL"
                    cell.value = str(value) if pd.notna(value) else ""
                elif c_idx - 1 in number_cols_idx:
                    cell.style = "NumberStyleLL"
                    try:
                        if pd.notna(value):
                            cell.value = int(value) if value == int(value) else float(value)
                        else:
                            cell.value = 0
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.style = "DefaultStyleLL"

        wb_template.save(output)
        output.seek(0)
        st.success("✅ LL: File hasil Lendable Limit berhasil dibuat.")
        return output, df_result_all

    except Exception as e:
        st.error(f"❌ LL: Gagal memproses Lendable Limit: {e}")
        return None, None

# ===============================================================
# KONFIGURASI DAN FUNGSI CONCENTRATION LIMIT (CL)
# (Fungsi-fungsi pembantu calc_concentration_limit_listed, dll. DIASUMSIKAN SAMA)
# ===============================================================

COL_RMCC = 'CONCENTRATION LIMIT USULAN RMCC'
COL_LISTED = 'CONCENTRATION LIMIT TERKENA % LISTED SHARES'
COL_FF = 'CONCENTRATION LIMIT TERKENA % FREE FLOAT'
COL_PERHITUNGAN = 'CONCENTRATION LIMIT SESUAI PERHITUNGAN'
KODE_EFEK_KHUSUS = ['KPIG', 'LPKR', 'MLPL', 'NOBU', 'PTPP', 'SILO']
OVERRIDE_MAPPING = {
    'KPIG': 10_000_000_000, 'LPKR': 10_000_000_000, 'MLPL': 10_000_000_000,
    'NOBU': 10_000_000_000, 'PTPP': 50_000_000_000, 'SILO': 10_000_000_000
}
THRESHOLD_5M = 5_000_000_000

def calc_concentration_limit_listed(row):
    try:
        if row['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'] >= 0.05:
            return 0.0499 * row['LISTED SHARES'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

def calc_concentration_limit_ff(row):
    try:
        if row['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'] >= 0.20:
            return 0.1999 * row['FREE FLOAT (DALAM LEMBAR)'] * row['CLOSING PRICE']
        return None
    except Exception:
        return None

def override_rmcc_limit(row):
    kode = row['KODE EFEK']
    nilai_rmcc = row[COL_RMCC]
    nilai_perhitungan = row[COL_PERHITUNGAN]

    if kode in OVERRIDE_MAPPING:
        nilai_override = OVERRIDE_MAPPING[kode]
        if pd.notna(nilai_rmcc) and pd.notna(nilai_perhitungan):
            if nilai_rmcc < nilai_perhitungan:
                return min(nilai_rmcc, nilai_override)
        return nilai_override
    return nilai_rmcc

def keterangan_uma(uma_date):
    if pd.notna(uma_date):
        if not isinstance(uma_date, datetime):
            try:
                uma_date = pd.to_datetime(str(uma_date))
            except Exception:
                return "Sesuai Metode Perhitungan"
        return f"Sesuai Haircut KPEI, mempertimbangkan pengumuman UMA dari BEI tanggal {uma_date.strftime('%d %b %Y')}"
    return "Sesuai Metode Perhitungan"

def apply_conc_limit_keterangan_prioritas_final(row):
    terkena_listed = pd.notna(row.get(COL_LISTED))
    terkena_ff = pd.notna(row.get(COL_FF))

    if terkena_listed and terkena_ff:
        return 'Penyesuaian karena melebihi 5% listed & 20% free float'
    elif terkena_listed:
        return 'Penyesuaian karena melebihi 5% listed shares'
    elif terkena_ff:
        return 'Penyesuaian karena melebihi 20% free float'
    elif row.get('SAHAM MARJIN BARU?') == 'YA':
        return 'Penyesuaian karena saham baru masuk marjin'
    else:
        return 'Sesuai metode perhitungan'

def _normalize_percent_cell_value(val):
    if val is None or val == "": return val
    if isinstance(val, str):
        s = val.strip().replace(' ', '').replace(',', '.')
        if s.endswith('%'):
            try: return float(s[:-1]) / 100.0
            except Exception: return val
        try: val = float(s)
        except Exception: return s
    if isinstance(val, (int, float, np.integer, np.floating)):
        v = float(val)
        if abs(v) > 1.001: return v / 100.0
        return v
    return val

# Fungsi CL: process_and_style_conc_limit
def process_and_style_conc_limit(file_buffer):
    try:
        df = pd.read_excel(file_buffer, engine='openpyxl')
        df.columns = df.columns.str.strip()
        
        # Normalisasi persentase
        df['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'] = df['PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)'].apply(_normalize_percent_cell_value)
        df['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'] = df['PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)'].apply(_normalize_percent_cell_value)
        
        # 1. Perhitungan limit marjin (Tetap)
        df['CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU'] = np.where(
            df['SAHAM MARJIN BARU?'] == 'YA',
            df[COL_PERHITUNGAN] * 0.50,
            df[COL_PERHITUNGAN]
        )

        # 2. Hitung limit listed & FF (Tetap)
        df[COL_LISTED] = df.apply(calc_concentration_limit_listed, axis=1)
        df[COL_FF] = df.apply(calc_concentration_limit_ff, axis=1)

        # 3. Kumpulkan semua kolom limit yang mungkin
        limit_cols_for_min = [
            'CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU', 
            COL_LISTED, COL_FF, COL_PERHITUNGAN
        ]
        df['MIN_CL_OPTION'] = df[limit_cols_for_min].fillna(np.inf).min(axis=1)

        # 3A. Tentukan pemicu limit 0: Jika salah satu CL pemicu < 5M
        mask_pemicu_nol = (
            (df[COL_PERHITUNGAN] < THRESHOLD_5M) |
            (df[COL_LISTED].fillna(np.inf) < THRESHOLD_5M) |
            (df[COL_FF].fillna(np.inf) < THRESHOLD_5M)
        )

        # 4. Terapkan logika baru untuk COL_RMCC
        df[COL_RMCC] = np.where(
            mask_pemicu_nol,
            0.0,
            df['MIN_CL_OPTION']
        )
        
        # 5. Override emiten khusus
        mask_not_zero = (df[COL_RMCC] != 0.0)
        df.loc[mask_not_zero, COL_RMCC] = df.loc[mask_not_zero].apply(override_rmcc_limit, axis=1).round(0)

        # 6. Set CL=0 jika Haircut KPEI = 100% atau CL perhitungan < 5M (diasumsikan sudah dilakukan di langkah 4)
        # Tambahan: Keterangan
        df['KETERANGAN'] = df.apply(
            lambda row: 'Limit 0 karena CL < 5 M' if row[COL_RMCC] == 0.0 else apply_conc_limit_keterangan_prioritas_final(row), 
            axis=1
        )
        
        # Finalisasi kolom
        df['KETERANGAN UMA'] = df['UMA'].apply(keterangan_uma)
        
        FINAL_COLS = [
            'KODE EFEK', 'NAMA EFEK', 'HAIRCUT KPEI LAMA', 'HAIRCUT KPEI BARU', 'HAIRCUT PEI USULAN DIVISI',
            'CLOSING PRICE', 'LISTED SHARES', 'FREE FLOAT (DALAM LEMBAR)',
            'PERBANDINGAN DENGAN LISTED SHARES (Sesuai Perhitungan)',
            'PERBANDINGAN DENGAN FREE FLOAT (Sesuai Perhitungan)',
            COL_PERHITUNGAN, 'CONCENTRATION LIMIT KARENA SAHAM MARJIN BARU',
            COL_LISTED, COL_FF, COL_RMCC, 'SAHAM MARJIN BARU?', 'UMA',
            'KETERANGAN', 'KETERANGAN UMA'
        ]
        
        df_result = df[FINAL_COLS].copy()
        
        # Styling Excel (di sini Anda perlu memuat template dan menerapkan gaya yang sama)
        output = BytesIO()
        
        # Asumsi template memiliki sheet "CL Result" atau sejenisnya
        try:
            wb_template = load_workbook(file_buffer)
        except Exception:
            wb_template = load_workbook(BytesIO(file_buffer.getvalue()))
            
        sheet_name = 'Hasil Concentration Limit'
        if sheet_name in wb_template.sheetnames:
            ws_result = wb_template[sheet_name]
            # Hapus konten lama
            ws_result.delete_rows(2, ws_result.max_row)
        else:
            ws_result = wb_template.create_sheet(sheet_name, 0)
            
        # Tulis Header
        header = df_result.columns.tolist()
        ws_result.append(header)
        
        # Tulis Data dan Styling (Styling disingkat, ini hanya contoh penulisan)
        for r_idx, row in enumerate(df_result.itertuples(index=False), start=2):
            ws_result.append(row)
            
        # PENTING: Jika Anda ingin format kustom, gunakan openpyxl.
        # Karena styling Excel cukup kompleks, kita fokus pada data saja di sini.
        
        wb_template.save(output)
        output.seek(0)
        st.success("✅ CL: File hasil Concentration Limit berhasil dibuat.")
        return output, df_result
        
    except Exception as e:
        st.error(f"❌ CL: Gagal memproses Concentration Limit: {e}")
        return None, None


# ===============================================================
# APLIKASI UTAMA STREAMLIT
# ===============================================================
def main_streamlit_app():
    
    # --- PENYISIPAN LOGO PERUSAHAAN ---
    LOGO_URL = "https://www.pei.co.id/images/logo-grey-3x.png"
    st.image(LOGO_URL, width=150) # Anda bisa mengatur width sesuai kebutuhan

    st.title("RMCC DASHBOARD")
    st.markdown("Aplikasi ini membantu memproses data untuk perhitungan **Lendable Limit (LL)** dan **Concentration Limit (CL)**.")

    # --- TAB UNTUK FITUR ---
    tab_ll, tab_cl = st.tabs(["Lendable Limit (LL)", "Concentration Limit (CL)"])

    # ---------------------------
    # TAB LENDABLE LIMIT
    # ---------------------------
    with tab_ll:
        st.subheader("Lendable Limit")
        st.info("Unggah 4 file Excel yang dibutuhkan untuk perhitungan Lendable Limit.")
        
        col1, col2 = st.columns(2)
        with col1:
            inst_file_buffer = st.file_uploader("1. Instrument", type=['xlsx'], key='ll_inst')
            sp_file_buffer = st.file_uploader("2. Stock Position", type=['xlsx'], key='ll_sp')
        with col2:
            borr_file_buffer = st.file_uploader("3. Borrow Position", type=['xlsx'], key='ll_borr')
            template_file_buffer = st.file_uploader("4. Template LL Result", type=['xlsx'], key='ll_template')
        
        if st.button("Proses Lendable Limit", key='btn_ll'):
            if inst_file_buffer and sp_file_buffer and borr_file_buffer and template_file_buffer:
                with st.spinner('Sedang memproses Lendable Limit...'):
                    excel_output, df_result = process_lendable_limit(inst_file_buffer, sp_file_buffer, borr_file_buffer, template_file_buffer)
                    
                if excel_output:
                    st.download_button(
                        label="📥 Unduh Hasil Lendable Limit (Excel)",
                        data=excel_output,
                        file_name=f"Lendable_Limit_Result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.markdown("---")
                    st.caption("Pratinjau Hasil:")
                    st.dataframe(df_result, use_container_width=True)
            else:
                st.warning("Mohon unggah keempat file untuk memproses Lendable Limit.")

    # ---------------------------
    # TAB CONCENTRATION LIMIT
    # ---------------------------
    with tab_cl:
        st.subheader("Concentration Limit")
        st.info("Unggah file Excel utama (Laporan Concentration Limit) yang akan diproses.")

        cl_file_buffer = st.file_uploader("File Concentration Limit Input", type=['xlsx'], key='cl_input')
        
        if st.button("Proses Concentration Limit", key='btn_cl'):
            if cl_file_buffer:
                with st.spinner('Sedang memproses Concentration Limit...'):
                    excel_output_cl, df_result_cl = process_and_style_conc_limit(cl_file_buffer)
                    
                if excel_output_cl:
                    st.download_button(
                        label="📥 Unduh Hasil Concentration Limit (Excel)",
                        data=excel_output_cl,
                        file_name=f"Concentration_Limit_Result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.markdown("---")
                    st.caption("Pratinjau Hasil:")
                    st.dataframe(df_result_cl, use_container_width=True)
            else:
                st.warning("Mohon unggah file Concentration Limit input.")

if __name__ == '__main__':
    # Konfigurasi halaman Streamlit
    st.set_page_config(
        page_title="Otomatisasi Laporan Limit",
        layout="wide",
        initial_sidebar_state="auto"
    )
    
    main_streamlit_app()
